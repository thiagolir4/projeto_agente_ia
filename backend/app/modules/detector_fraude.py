# -*- coding: utf-8 -*-
"""
Módulo de Detecção de Fraude
Implementa algoritmos para identificar indícios de fraude em dados de vendas, devoluções e ajustes.
"""

from datetime import datetime, timedelta
from typing import List, Dict, Any, Tuple, Optional
from pymongo import MongoClient
from collections import defaultdict, Counter
import pandas as pd
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io


class DetectorFraude:
    """
    Classe principal para detecção de fraudes em dados de vendas e movimentações.
    """
    
    def __init__(self, mongo_client: MongoClient, database_name: str):
        """
        Inicializa o detector de fraude.
        
        Args:
            mongo_client: Cliente MongoDB conectado
            database_name: Nome do banco de dados
        """
        self.client = mongo_client
        self.db = mongo_client[database_name]
        self.resultados_fraude = []
        
        # Configurações para detecção de fraude
        self.config = {
            'percentual_troca_suspeito': 0.15,  # 15% de trocas é suspeito
            'percentual_cancelamento_suspeito': 0.10,  # 10% de cancelamentos é suspeito
            'intervalo_suspeito_horas': 24,  # Movimentações no mesmo dia são suspeitas
            'minimo_ocorrencias_reincidencia': 3,  # Mínimo de ocorrências para ser reincidente
            'diferenca_valor_suspeita': 0.20  # 20% de diferença de valor é suspeita
        }
    
    def detectar_trocas_mais_saidas_que_entradas(self) -> List[Dict[str, Any]]:
        """
        Detecta produtos com volume anômalo de devoluções/trocas.
        Analisa se um produto tem mais devoluções que o esperado.
        
        Returns:
            Lista de suspeitas de fraude encontradas
        """
        print("Analisando volume anômalo de devoluções...")
        suspeitas = []
        
        try:
            # Buscar dados de devolução
            devolucoes = list(self.db.DEVOLUCAO.find({}))
            
            if not devolucoes:
                print(" Nenhuma devolução encontrada para análise")
                return []
            
            # Agrupar por SKU e loja
            movimentacoes = defaultdict(lambda: {'total_devolucoes': 0, 'detalhes': []})
            
            for dev in devolucoes:
                sku = dev.get('SKU', '')
                loja = dev.get('LOJA', '')
                tipo = dev.get('TIPOMOVIMENTACAO', '')
                valor = self._converter_valor(dev.get('DIFERENCA_VALOR', '0'))
                
                chave = f"{sku}_{loja}"
                
                # Contar todas as devoluções (independente do tipo)
                movimentacoes[chave]['total_devolucoes'] += 1
                
                movimentacoes[chave]['detalhes'].append({
                    'data': dev.get('DATA_DEVOLUCAO', ''),
                    'tipo': tipo,
                    'valor': valor,
                    'usuario': dev.get('IDUSUARIO', ''),
                    'id_devolucao': dev.get('ID_DEVOLUCAO', '')
                })
            
            # Calcular estatísticas gerais
            total_devolucoes = len(devolucoes)
            total_produtos = len(movimentacoes)
            media_devolucoes_por_produto = total_devolucoes / max(total_produtos, 1)
            
            # Identificar suspeitas (produtos com devoluções muito acima da média)
            limite_suspeito = media_devolucoes_por_produto * 3  # 3x a média
            limite_alto_risco = media_devolucoes_por_produto * 5  # 5x a média
            
            for chave, dados in movimentacoes.items():
                total_dev = dados['total_devolucoes']
                
                if total_dev > limite_suspeito:
                    sku, loja = chave.split('_', 1)
                    
                    # Calcular percentual acima da média
                    percentual_acima_media = round((total_dev / media_devolucoes_por_produto - 1) * 100, 2)
                    
                    suspeita = {
                        'tipo_fraude': 'Volume anômalo de devoluções',
                        'nivel_risco': 'ALTO' if total_dev > limite_alto_risco else 'MÉDIO',
                        'sku': sku,
                        'loja': loja,
                        'total_devolucoes': total_dev,
                        'media_geral': round(media_devolucoes_por_produto, 2),
                        'percentual_acima_media': percentual_acima_media,
                        'detalhes': dados['detalhes'],
                        'timestamp_analise': datetime.now().isoformat()
                    }
                    suspeitas.append(suspeita)
            
            print(f" Encontradas {len(suspeitas)} suspeitas de volume anômalo de devoluções")
            print(f" Estatísticas: {total_devolucoes} devoluções, {total_produtos} produtos, média: {media_devolucoes_por_produto:.2f}")
            return suspeitas
            
        except Exception as e:
            print(f" Erro ao analisar devoluções: {e}")
            return []
    
    def detectar_produtos_ajustes_trocas_simultaneos(self) -> List[Dict[str, Any]]:
        """
        Detecta produtos que aparecem simultaneamente em ajustes e trocas.
        
        Returns:
            Lista de suspeitas de fraude encontradas
        """
        print(" Analisando produtos em ajustes e trocas simultâneos...")
        suspeitas = []
        
        try:
            # Buscar SKUs de ajustes de estoque
            ajustes_skus = set()
            ajustes_detalhes = {}
            
            if 'AJUSTES ESTOQUE' in self.db.list_collection_names():
                ajustes = list(self.db['AJUSTES ESTOQUE'].find({}))
                for ajuste in ajustes:
                    sku = ajuste.get('SKU', '')
                    if sku:
                        ajustes_skus.add(sku)
                        ajustes_detalhes[sku] = ajuste
            
            # Buscar SKUs de devoluções/trocas
            devolucoes_skus = set()
            devolucoes_detalhes = {}
            
            devolucoes = list(self.db.DEVOLUCAO.find({}))
            for dev in devolucoes:
                sku = dev.get('SKU', '')
                if sku:
                    devolucoes_skus.add(sku)
                    if sku not in devolucoes_detalhes:
                        devolucoes_detalhes[sku] = []
                    devolucoes_detalhes[sku].append(dev)
            
            # Encontrar interseção
            skus_simultaneos = ajustes_skus.intersection(devolucoes_skus)
            
            for sku in skus_simultaneos:
                # Analisar datas para verificar se são próximas
                ajuste_data = ajustes_detalhes[sku].get('DATA_AJUSTE', '')
                devolucoes_sku = devolucoes_detalhes[sku]
                
                for dev in devolucoes_sku:
                    dev_data = dev.get('DATA_DEVOLUCAO', '')
                    
                    # Verificar se as datas são próximas (mesmo dia ou próximo)
                    if self._datas_proximas(ajuste_data, dev_data):
                        suspeita = {
                            'tipo_fraude': 'Produto em ajuste e troca simultâneos',
                            'nivel_risco': 'ALTO',
                            'sku': sku,
                            'loja_ajuste': ajustes_detalhes[sku].get('LOJA', ''),
                            'loja_devolucao': dev.get('LOJA', ''),
                            'data_ajuste': ajuste_data,
                            'data_devolucao': dev_data,
                            'detalhes_ajuste': ajustes_detalhes[sku],
                            'detalhes_devolucao': dev,
                            'timestamp_analise': datetime.now().isoformat()
                        }
                        suspeitas.append(suspeita)
            
            print(f" Encontradas {len(suspeitas)} suspeitas de produtos em ajustes e trocas simultâneos")
            return suspeitas
            
        except Exception as e:
            print(f" Erro ao analisar ajustes e trocas simultâneos: {e}")
            return []
    
    def detectar_percentuais_suspeitos_trocas_cancelamentos(self) -> List[Dict[str, Any]]:
        """
        Detecta percentuais de trocas e cancelamentos acima do padrão.
        
        Returns:
            Lista de suspeitas de fraude encontradas
        """
        print(" Analisando percentuais suspeitos de trocas e cancelamentos...")
        suspeitas = []
        
        try:
            # Analisar por loja
            lojas_stats = self._calcular_estatisticas_lojas()
            
            for loja, stats in lojas_stats.items():
                # Verificar percentual de trocas
                if stats['percentual_trocas'] > self.config['percentual_troca_suspeito']:
                    suspeita = {
                        'tipo_fraude': 'Percentual de trocas acima do padrão',
                        'nivel_risco': 'ALTO' if stats['percentual_trocas'] > 0.25 else 'MÉDIO',
                        'loja': loja,
                        'percentual_trocas': round(stats['percentual_trocas'] * 100, 2),
                        'limite_suspeito': round(self.config['percentual_troca_suspeito'] * 100, 2),
                        'total_movimentacoes': stats['total_movimentacoes'],
                        'total_trocas': stats['total_trocas'],
                        'detalhes': stats,
                        'timestamp_analise': datetime.now().isoformat()
                    }
                    suspeitas.append(suspeita)
                
                # Verificar percentual de cancelamentos
                if stats['percentual_cancelamentos'] > self.config['percentual_cancelamento_suspeito']:
                    suspeita = {
                        'tipo_fraude': 'Percentual de cancelamentos acima do padrão',
                        'nivel_risco': 'ALTO' if stats['percentual_cancelamentos'] > 0.15 else 'MÉDIO',
                        'loja': loja,
                        'percentual_cancelamentos': round(stats['percentual_cancelamentos'] * 100, 2),
                        'limite_suspeito': round(self.config['percentual_cancelamento_suspeito'] * 100, 2),
                        'total_movimentacoes': stats['total_movimentacoes'],
                        'total_cancelamentos': stats['total_cancelamentos'],
                        'detalhes': stats,
                        'timestamp_analise': datetime.now().isoformat()
                    }
                    suspeitas.append(suspeita)
            
            print(f" Encontradas {len(suspeitas)} suspeitas de percentuais anômalos")
            return suspeitas
            
        except Exception as e:
            print(f" Erro ao analisar percentuais: {e}")
            return []
    
    def detectar_movimentacoes_suspeitas_curto_intervalo(self) -> List[Dict[str, Any]]:
        """
        Detecta produtos com movimentações suspeitas em curto intervalo.
        
        Returns:
            Lista de suspeitas de fraude encontradas
        """
        print(" Analisando movimentações suspeitas em curto intervalo...")
        suspeitas = []
        
        try:
            # Buscar todas as movimentações ordenadas por data
            todas_movimentacoes = []
            
            # Devoluções
            devolucoes = list(self.db.DEVOLUCAO.find({}))
            for dev in devolucoes:
                todas_movimentacoes.append({
                    'tipo': 'DEVOLUCAO',
                    'sku': dev.get('SKU', ''),
                    'loja': dev.get('LOJA', ''),
                    'usuario': dev.get('IDUSUARIO', ''),
                    'data': dev.get('DATA_DEVOLUCAO', ''),
                    'tipo_movimentacao': dev.get('TIPOMOVIMENTACAO', ''),
                    'valor': self._converter_valor(dev.get('DIFERENCA_VALOR', '0')),
                    'documento': dev
                })
            
            # Ajustes
            if 'AJUSTES ESTOQUE' in self.db.list_collection_names():
                ajustes = list(self.db['AJUSTES ESTOQUE'].find({}))
                for ajuste in ajustes:
                    todas_movimentacoes.append({
                        'tipo': 'AJUSTE',
                        'sku': ajuste.get('SKU', ''),
                        'loja': ajuste.get('LOJA', ''),
                        'usuario': ajuste.get('IDUSUARIO', ''),
                        'data': ajuste.get('DATA_AJUSTE', ''),
                        'tipo_movimentacao': ajuste.get('TIPO_AJUSTE', ''),
                        'valor': self._converter_valor(ajuste.get('DIFERENCA_VALOR', '0')),
                        'documento': ajuste
                    })
            
            # Cancelamentos
            if 'CANCELAMENTO' in self.db.list_collection_names():
                cancelamentos = list(self.db.CANCELAMENTO.find({}))
                for cancel in cancelamentos:
                    todas_movimentacoes.append({
                        'tipo': 'CANCELAMENTO',
                        'sku': cancel.get('SKU', ''),
                        'loja': cancel.get('LOJA', ''),
                        'usuario': cancel.get('IDUSUARIO', ''),
                        'data': cancel.get('DATA_CANCELAMENTO', ''),
                        'tipo_movimentacao': 'CANCELAMENTO',
                        'valor': self._converter_valor(cancel.get('VALOR_CANCELAMENTO', '0')),
                        'documento': cancel
                    })
            
            # Agrupar por SKU e loja
            movimentacoes_por_sku = defaultdict(list)
            for mov in todas_movimentacoes:
                if mov['sku'] and mov['loja']:
                    chave = f"{mov['sku']}_{mov['loja']}"
                    movimentacoes_por_sku[chave].append(mov)
            
            # Analisar intervalos suspeitos
            for chave, movimentacoes in movimentacoes_por_sku.items():
                if len(movimentacoes) < 2:
                    continue
                
                # Ordenar por data
                movimentacoes.sort(key=lambda x: x['data'])
                
                for i in range(len(movimentacoes) - 1):
                    mov1 = movimentacoes[i]
                    mov2 = movimentacoes[i + 1]
                    
                    if self._datas_proximas(mov1['data'], mov2['data'], horas=self.config['intervalo_suspeito_horas']):
                        sku, loja = chave.split('_', 1)
                        
                        suspeita = {
                            'tipo_fraude': 'Movimentações suspeitas em curto intervalo',
                            'nivel_risco': 'ALTO',
                            'sku': sku,
                            'loja': loja,
                            'usuario': mov1['usuario'],
                            'data_movimentacao_1': mov1['data'],
                            'data_movimentacao_2': mov2['data'],
                            'tipo_movimentacao_1': mov1['tipo_movimentacao'],
                            'tipo_movimentacao_2': mov2['tipo_movimentacao'],
                            'valor_1': mov1['valor'],
                            'valor_2': mov2['valor'],
                            'detalhes_movimentacao_1': mov1['documento'],
                            'detalhes_movimentacao_2': mov2['documento'],
                            'timestamp_analise': datetime.now().isoformat()
                        }
                        suspeitas.append(suspeita)
            
            print(f" Encontradas {len(suspeitas)} suspeitas de movimentações em curto intervalo")
            return suspeitas
            
        except Exception as e:
            print(f" Erro ao analisar movimentações em curto intervalo: {e}")
            return []
    
    def detectar_clientes_produtos_reincidentes(self) -> List[Dict[str, Any]]:
        """
        Detecta clientes e produtos reincidentes em trocas ou divergências.
        
        Returns:
            Lista de suspeitas de fraude encontradas
        """
        print(" Analisando clientes e produtos reincidentes...")
        suspeitas = []
        
        try:
            # Analisar clientes reincidentes
            clientes_stats = self._calcular_estatisticas_clientes()
            
            for cliente, stats in clientes_stats.items():
                if stats['total_ocorrencias'] >= self.config['minimo_ocorrencias_reincidencia']:
                    suspeita = {
                        'tipo_fraude': 'Cliente reincidente em trocas/divergências',
                        'nivel_risco': 'ALTO' if stats['total_ocorrencias'] > 10 else 'MÉDIO',
                        'cliente': cliente,
                        'total_ocorrencias': stats['total_ocorrencias'],
                        'total_trocas': stats['total_trocas'],
                        'total_cancelamentos': stats['total_cancelamentos'],
                        'total_ajustes': stats['total_ajustes'],
                        'lojas_envolvidas': list(stats['lojas_envolvidas']),
                        'skus_envolvidos': list(stats['skus_envolvidos']),
                        'detalhes': stats,
                        'timestamp_analise': datetime.now().isoformat()
                    }
                    suspeitas.append(suspeita)
            
            # Analisar produtos reincidentes
            produtos_stats = self._calcular_estatisticas_produtos()
            
            for produto, stats in produtos_stats.items():
                if stats['total_ocorrencias'] >= self.config['minimo_ocorrencias_reincidencia']:
                    suspeita = {
                        'tipo_fraude': 'Produto reincidente em trocas/divergências',
                        'nivel_risco': 'ALTO' if stats['total_ocorrencias'] > 15 else 'MÉDIO',
                        'produto': produto,
                        'total_ocorrencias': stats['total_ocorrencias'],
                        'total_trocas': stats['total_trocas'],
                        'total_cancelamentos': stats['total_cancelamentos'],
                        'total_ajustes': stats['total_ajustes'],
                        'lojas_envolvidas': list(stats['lojas_envolvidas']),
                        'clientes_envolvidos': list(stats['clientes_envolvidos']),
                        'detalhes': stats,
                        'timestamp_analise': datetime.now().isoformat()
                    }
                    suspeitas.append(suspeita)
            
            print(f" Encontradas {len(suspeitas)} suspeitas de reincidência")
            return suspeitas
            
        except Exception as e:
            print(f" Erro ao analisar reincidências: {e}")
            return []
    
    def executar_analise_completa_fraude(self) -> Dict[str, Any]:
        """
        Executa análise completa de fraude com todos os algoritmos.
        
        Returns:
            Relatório completo de análise de fraude
        """
        print("Iniciando análise completa de detecção de fraude...")
        
        inicio_analise = datetime.now()
        todas_suspeitas = []
        
        # Executar todos os algoritmos de detecção
        algoritmos = [
            ('Trocas desbalanceadas', self.detectar_trocas_mais_saidas_que_entradas),
            ('Produtos em ajustes e trocas simultâneos', self.detectar_produtos_ajustes_trocas_simultaneos),
            ('Percentuais suspeitos', self.detectar_percentuais_suspeitos_trocas_cancelamentos),
            ('Movimentações em curto intervalo', self.detectar_movimentacoes_suspeitas_curto_intervalo),
            ('Reincidências', self.detectar_clientes_produtos_reincidentes)
        ]
        
        for nome_algoritmo, funcao_algoritmo in algoritmos:
            print(f"\n Executando: {nome_algoritmo}")
            try:
                suspeitas = funcao_algoritmo()
                todas_suspeitas.extend(suspeitas)
                print(f" {nome_algoritmo}: {len(suspeitas)} suspeitas encontradas")
            except Exception as e:
                print(f" Erro em {nome_algoritmo}: {e}")
        
        fim_analise = datetime.now()
        tempo_analise = (fim_analise - inicio_analise).total_seconds()
        
        # Gerar relatório
        relatorio = {
            'timestamp_analise': inicio_analise.isoformat(),
            'tempo_analise_segundos': round(tempo_analise, 2),
            'total_suspeitas': len(todas_suspeitas),
            'suspeitas_por_tipo': self._agrupar_suspeitas_por_tipo(todas_suspeitas),
            'suspeitas_por_nivel_risco': self._agrupar_suspeitas_por_risco(todas_suspeitas),
            'resumo_executivo': self._gerar_resumo_executivo(todas_suspeitas),
            'detalhes_suspeitas': todas_suspeitas,
            'configuracao_utilizada': self.config
        }
        
        print(f"\n Análise completa finalizada em {tempo_analise:.2f} segundos")
        print(f" Total de suspeitas encontradas: {len(todas_suspeitas)}")
        
        return relatorio
    
    def _converter_valor(self, valor_str: str) -> float:
        """Converte string de valor para float."""
        try:
            if not valor_str or valor_str == '':
                return 0.0
            # Remover vírgulas e converter para float
            valor_limpo = str(valor_str).replace(',', '.').replace('R$', '').replace(' ', '')
            return float(valor_limpo)
        except:
            return 0.0
    
    def _datas_proximas(self, data1: str, data2: str, horas: int = 24) -> bool:
        """Verifica se duas datas estão próximas dentro do intervalo especificado."""
        try:
            if not data1 or not data2:
                return False
            
            # Converter strings para datetime (assumindo formato DD/MM/YYYY)
            dt1 = datetime.strptime(data1, '%d/%m/%Y') if '/' in data1 else datetime.fromisoformat(data1)
            dt2 = datetime.strptime(data2, '%d/%m/%Y') if '/' in data2 else datetime.fromisoformat(data2)
            
            diferenca = abs((dt2 - dt1).total_seconds() / 3600)  # em horas
            return diferenca <= horas
        except:
            return False
    
    def _calcular_estatisticas_lojas(self) -> Dict[str, Dict[str, Any]]:
        """Calcula estatísticas por loja."""
        stats = defaultdict(lambda: {
            'total_movimentacoes': 0,
            'total_trocas': 0,
            'total_cancelamentos': 0,
            'percentual_trocas': 0.0,
            'percentual_cancelamentos': 0.0
        })
        
        # Analisar devoluções
        devolucoes = list(self.db.DEVOLUCAO.find({}))
        for dev in devolucoes:
            loja = dev.get('LOJA', '')
            if loja:
                stats[loja]['total_movimentacoes'] += 1
                if 'TROCA' in dev.get('TIPOMOVIMENTACAO', '').upper():
                    stats[loja]['total_trocas'] += 1
        
        # Analisar cancelamentos
        if 'CANCELAMENTO' in self.db.list_collection_names():
            cancelamentos = list(self.db.CANCELAMENTO.find({}))
            for cancel in cancelamentos:
                loja = cancel.get('LOJA', '')
                if loja:
                    stats[loja]['total_movimentacoes'] += 1
                    stats[loja]['total_cancelamentos'] += 1
        
        # Calcular percentuais
        for loja, dados in stats.items():
            if dados['total_movimentacoes'] > 0:
                dados['percentual_trocas'] = dados['total_trocas'] / dados['total_movimentacoes']
                dados['percentual_cancelamentos'] = dados['total_cancelamentos'] / dados['total_movimentacoes']
        
        return dict(stats)
    
    def _calcular_estatisticas_clientes(self) -> Dict[str, Dict[str, Any]]:
        """Calcula estatísticas por cliente."""
        stats = defaultdict(lambda: {
            'total_ocorrencias': 0,
            'total_trocas': 0,
            'total_cancelamentos': 0,
            'total_ajustes': 0,
            'lojas_envolvidas': set(),
            'skus_envolvidos': set()
        })
        
        # Analisar devoluções
        devolucoes = list(self.db.DEVOLUCAO.find({}))
        for dev in devolucoes:
            cliente = dev.get('IDUSUARIO', '')
            if cliente:
                stats[cliente]['total_ocorrencias'] += 1
                if 'TROCA' in dev.get('TIPOMOVIMENTACAO', '').upper():
                    stats[cliente]['total_trocas'] += 1
                stats[cliente]['lojas_envolvidas'].add(dev.get('LOJA', ''))
                stats[cliente]['skus_envolvidos'].add(dev.get('SKU', ''))
        
        # Analisar cancelamentos
        if 'CANCELAMENTO' in self.db.list_collection_names():
            cancelamentos = list(self.db.CANCELAMENTO.find({}))
            for cancel in cancelamentos:
                cliente = cancel.get('IDUSUARIO', '')
                if cliente:
                    stats[cliente]['total_ocorrencias'] += 1
                    stats[cliente]['total_cancelamentos'] += 1
                    stats[cliente]['lojas_envolvidas'].add(cancel.get('LOJA', ''))
                    stats[cliente]['skus_envolvidos'].add(cancel.get('SKU', ''))
        
        # Analisar ajustes
        if 'AJUSTES ESTOQUE' in self.db.list_collection_names():
            ajustes = list(self.db['AJUSTES ESTOQUE'].find({}))
            for ajuste in ajustes:
                cliente = ajuste.get('IDUSUARIO', '')
                if cliente:
                    stats[cliente]['total_ocorrencias'] += 1
                    stats[cliente]['total_ajustes'] += 1
                    stats[cliente]['lojas_envolvidas'].add(ajuste.get('LOJA', ''))
                    stats[cliente]['skus_envolvidos'].add(ajuste.get('SKU', ''))
        
        return dict(stats)
    
    def _calcular_estatisticas_produtos(self) -> Dict[str, Dict[str, Any]]:
        """Calcula estatísticas por produto."""
        stats = defaultdict(lambda: {
            'total_ocorrencias': 0,
            'total_trocas': 0,
            'total_cancelamentos': 0,
            'total_ajustes': 0,
            'lojas_envolvidas': set(),
            'clientes_envolvidos': set()
        })
        
        # Analisar devoluções
        devolucoes = list(self.db.DEVOLUCAO.find({}))
        for dev in devolucoes:
            sku = dev.get('SKU', '')
            if sku:
                stats[sku]['total_ocorrencias'] += 1
                if 'TROCA' in dev.get('TIPOMOVIMENTACAO', '').upper():
                    stats[sku]['total_trocas'] += 1
                stats[sku]['lojas_envolvidas'].add(dev.get('LOJA', ''))
                stats[sku]['clientes_envolvidos'].add(dev.get('IDUSUARIO', ''))
        
        # Analisar cancelamentos
        if 'CANCELAMENTO' in self.db.list_collection_names():
            cancelamentos = list(self.db.CANCELAMENTO.find({}))
            for cancel in cancelamentos:
                sku = cancel.get('SKU', '')
                if sku:
                    stats[sku]['total_ocorrencias'] += 1
                    stats[sku]['total_cancelamentos'] += 1
                    stats[sku]['lojas_envolvidas'].add(cancel.get('LOJA', ''))
                    stats[sku]['clientes_envolvidos'].add(cancel.get('IDUSUARIO', ''))
        
        # Analisar ajustes
        if 'AJUSTES ESTOQUE' in self.db.list_collection_names():
            ajustes = list(self.db['AJUSTES ESTOQUE'].find({}))
            for ajuste in ajustes:
                sku = ajuste.get('SKU', '')
                if sku:
                    stats[sku]['total_ocorrencias'] += 1
                    stats[sku]['total_ajustes'] += 1
                    stats[sku]['lojas_envolvidas'].add(ajuste.get('LOJA', ''))
                    stats[sku]['clientes_envolvidos'].add(ajuste.get('IDUSUARIO', ''))
        
        return dict(stats)
    
    def _agrupar_suspeitas_por_tipo(self, suspeitas: List[Dict[str, Any]]) -> Dict[str, int]:
        """Agrupa suspeitas por tipo de fraude."""
        contadores = Counter()
        for suspeita in suspeitas:
            contadores[suspeita.get('tipo_fraude', 'Desconhecido')] += 1
        return dict(contadores)
    
    def _agrupar_suspeitas_por_risco(self, suspeitas: List[Dict[str, Any]]) -> Dict[str, int]:
        """Agrupa suspeitas por nível de risco."""
        contadores = Counter()
        for suspeita in suspeitas:
            contadores[suspeita.get('nivel_risco', 'Desconhecido')] += 1
        return dict(contadores)
    
    def _gerar_resumo_executivo(self, suspeitas: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Gera resumo executivo da análise."""
        total_suspeitas = len(suspeitas)
        suspeitas_alto_risco = len([s for s in suspeitas if s.get('nivel_risco') == 'ALTO'])
        suspeitas_medio_risco = len([s for s in suspeitas if s.get('nivel_risco') == 'MÉDIO'])
        
        return {
            'total_suspeitas': total_suspeitas,
            'suspeitas_alto_risco': suspeitas_alto_risco,
            'suspeitas_medio_risco': suspeitas_medio_risco,
            'percentual_alto_risco': round(suspeitas_alto_risco / max(total_suspeitas, 1) * 100, 2),
            'recomendacoes': self._gerar_recomendacoes(suspeitas)
        }
    
    def _gerar_recomendacoes(self, suspeitas: List[Dict[str, Any]]) -> List[str]:
        """Gera recomendações baseadas nas suspeitas encontradas."""
        recomendacoes = []
        
        if not suspeitas:
            recomendacoes.append(" Nenhuma suspeita de fraude detectada. Sistema operando normalmente.")
            return recomendacoes
        
        suspeitas_alto_risco = [s for s in suspeitas if s.get('nivel_risco') == 'ALTO']
        
        if suspeitas_alto_risco:
            recomendacoes.append("ATENÇÃO: Suspeitas de ALTO RISCO detectadas. Investigação imediata recomendada.")
        
        tipos_fraude = set(s.get('tipo_fraude', '') for s in suspeitas)
        
        if 'Trocas com mais saídas que entradas' in tipos_fraude:
            recomendacoes.append(" Revisar processos de controle de estoque para trocas desbalanceadas.")
        
        if 'Produto em ajuste e troca simultâneos' in tipos_fraude:
            recomendacoes.append(" Implementar validação cruzada entre ajustes de estoque e trocas.")
        
        if 'Percentual de trocas acima do padrão' in tipos_fraude:
            recomendacoes.append(" Investigar lojas com percentuais anômalos de trocas.")
        
        if 'Movimentações suspeitas em curto intervalo' in tipos_fraude:
            recomendacoes.append("⏰ Implementar cooldown entre movimentações do mesmo produto/usuário.")
        
        if 'Cliente reincidente' in tipos_fraude or 'Produto reincidente' in tipos_fraude:
            recomendacoes.append(" Revisar políticas para clientes e produtos com histórico de problemas.")
        
        return recomendacoes
    
    def gerar_excel_relatorio_fraude(self, relatorio: Dict[str, Any]) -> io.BytesIO:
        """
        Gera arquivo Excel do relatório de análise de fraude.
        
        Args:
            relatorio: Dicionário com dados do relatório de fraude
            
        Returns:
            BytesIO com o arquivo Excel
        """
        try:
            # Criar workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Relatório de Fraude"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF", size=14)
            subheader_font = Font(bold=True, color="000000", size=12)
            normal_font = Font(size=11)
            
            header_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
            subheader_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            high_risk_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
            medium_risk_fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
            
            center_alignment = Alignment(horizontal="center", vertical="center")
            left_alignment = Alignment(horizontal="left", vertical="center")
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            row = 1
            
            # Cabeçalho principal
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = "RELATÓRIO DE ANÁLISE DE FRAUDE"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].alignment = center_alignment
            row += 1
            
            # Data e tempo
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = f"Análise executada em: {relatorio.get('timestamp_analise', 'N/A')}"
            ws[f'A{row}'].font = normal_font
            ws[f'A{row}'].alignment = center_alignment
            row += 1
            
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = f"Tempo de análise: {relatorio.get('tempo_analise_segundos', 0)} segundos"
            ws[f'A{row}'].font = normal_font
            ws[f'A{row}'].alignment = center_alignment
            row += 2
            
            # Resumo executivo
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = "RESUMO EXECUTIVO"
            ws[f'A{row}'].font = subheader_font
            ws[f'A{row}'].fill = subheader_fill
            ws[f'A{row}'].alignment = left_alignment
            row += 1
            
            # Estatísticas
            total_suspeitas = relatorio.get('total_suspeitas', 0)
            suspeitas_por_risco = relatorio.get('suspeitas_por_nivel_risco', {})
            resumo_executivo = relatorio.get('resumo_executivo', {})
            
            stats_data = [
                ["Métrica", "Valor"],
                ["Total de Suspeitas", total_suspeitas],
                ["Alto Risco", suspeitas_por_risco.get('ALTO', 0)],
                ["Médio Risco", suspeitas_por_risco.get('MÉDIO', 0)],
                ["% Alto Risco", f"{resumo_executivo.get('percentual_alto_risco', 0)}%"]
            ]
            
            for stat_row in stats_data:
                ws[f'A{row}'] = stat_row[0]
                ws[f'B{row}'] = stat_row[1]
                ws[f'A{row}'].font = normal_font
                ws[f'B{row}'].font = normal_font
                ws[f'A{row}'].border = thin_border
                ws[f'B{row}'].border = thin_border
                row += 1
            
            row += 1
            
            # Tipos de fraude detectados
            suspeitas_por_tipo = relatorio.get('suspeitas_por_tipo', {})
            if suspeitas_por_tipo:
                ws.merge_cells(f'A{row}:F{row}')
                ws[f'A{row}'] = "TIPOS DE FRAUDE DETECTADOS"
                ws[f'A{row}'].font = subheader_font
                ws[f'A{row}'].fill = subheader_fill
                ws[f'A{row}'].alignment = left_alignment
                row += 1
                
                ws[f'A{row}'] = "Tipo de Fraude"
                ws[f'B{row}'] = "Quantidade"
                ws[f'A{row}'].font = subheader_font
                ws[f'B{row}'].font = subheader_font
                ws[f'A{row}'].border = thin_border
                ws[f'B{row}'].border = thin_border
                row += 1
                
                for tipo, quantidade in suspeitas_por_tipo.items():
                    ws[f'A{row}'] = tipo
                    ws[f'B{row}'] = quantidade
                    ws[f'A{row}'].font = normal_font
                    ws[f'B{row}'].font = normal_font
                    ws[f'A{row}'].border = thin_border
                    ws[f'B{row}'].border = thin_border
                    row += 1
                
                row += 1
            
            
            # Detalhes das suspeitas
            detalhes_suspeitas = relatorio.get('detalhes_suspeitas', [])
            if detalhes_suspeitas:
                ws.merge_cells(f'A{row}:F{row}')
                ws[f'A{row}'] = "DETALHES DAS SUSPEITAS"
                ws[f'A{row}'].font = subheader_font
                ws[f'A{row}'].fill = subheader_fill
                ws[f'A{row}'].alignment = left_alignment
                row += 1
                
                # Cabeçalho da tabela
                headers = ["Tipo de Fraude", "Nível de Risco", "SKU", "Loja", "Cliente", "Detalhes"]
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = subheader_font
                    cell.border = thin_border
                    cell.alignment = center_alignment
                row += 1
                
                # Dados das suspeitas - TODAS as suspeitas incluídas
                print(f" Incluindo todas as {len(detalhes_suspeitas)} suspeitas no Excel...")
                for i, suspeita in enumerate(detalhes_suspeitas):
                    # Extrair dados da suspeita
                    tipo_fraude = suspeita.get('tipo_fraude', 'N/A')
                    nivel_risco = suspeita.get('nivel_risco', 'N/A')
                    sku = suspeita.get('sku', 'N/A')
                    loja = suspeita.get('loja', 'N/A')
                    cliente = suspeita.get('cliente', 'N/A')
                    
                    # Criar detalhes completos (sem truncamento)
                    detalhes_resumo = []
                    for chave, valor in suspeita.items():
                        if chave not in ['tipo_fraude', 'nivel_risco', 'sku', 'loja', 'cliente', 'timestamp_analise'] and valor:
                            if isinstance(valor, list):
                                # Incluir TODOS os valores da lista, sem truncamento
                                detalhes_resumo.append(f"{chave}: {', '.join(map(str, valor))}")
                            else:
                                detalhes_resumo.append(f"{chave}: {valor}")
                    
                    # Incluir TODOS os detalhes, sem limitação
                    detalhes_texto = " | ".join(detalhes_resumo) if detalhes_resumo else "N/A"
                    
                    # Adicionar linha
                    data_row = [tipo_fraude, nivel_risco, sku, loja, cliente, detalhes_texto]
                    for col, value in enumerate(data_row, 1):
                        cell = ws.cell(row=row, column=col, value=value)
                        cell.font = normal_font
                        cell.border = thin_border
                        if col == 6:  # Coluna de detalhes - quebra de linha
                            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                        else:
                            cell.alignment = left_alignment
                        
                        # Aplicar cor baseada no nível de risco
                        if col == 2:  # Coluna do nível de risco
                            if nivel_risco == "ALTO":
                                cell.fill = high_risk_fill
                                cell.font = Font(bold=True, color="FFFFFF", size=11)
                            elif nivel_risco == "MÉDIO":
                                cell.fill = medium_risk_fill
                                cell.font = Font(bold=True, color="000000", size=11)
                    
                    row += 1
                    
                    # Log de progresso a cada 1000 registros
                    if (i + 1) % 1000 == 0:
                        print(f" Processadas {i + 1} suspeitas...")
                
                # Adicionar informações finais
                ws.merge_cells(f'A{row}:F{row}')
                ws[f'A{row}'] = f"Total de suspeitas incluídas: {len(detalhes_suspeitas)}"
                ws[f'A{row}'].font = Font(bold=True, color="495057")
                ws[f'A{row}'].alignment = center_alignment
                row += 1
            
            # Ajustar largura das colunas
            for col in range(1, 7):
                column_letter = get_column_letter(col)
                if col == 1:  # Tipo de Fraude
                    ws.column_dimensions[column_letter].width = 30
                elif col == 2:  # Nível de Risco
                    ws.column_dimensions[column_letter].width = 15
                elif col == 3:  # SKU
                    ws.column_dimensions[column_letter].width = 20
                elif col == 4:  # Loja
                    ws.column_dimensions[column_letter].width = 10
                elif col == 5:  # Cliente
                    ws.column_dimensions[column_letter].width = 20
                elif col == 6:  # Detalhes
                    ws.column_dimensions[column_letter].width = 100
            
            # Salvar em BytesIO
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            
            return excel_buffer
            
        except Exception as e:
            print(f" Erro ao gerar Excel: {e}")
            # Retornar Excel vazio em caso de erro
            wb = Workbook()
            ws = wb.active
            ws['A1'] = f"Erro ao gerar relatório: {str(e)}"
            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)
            return excel_buffer
